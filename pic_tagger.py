import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import Scrollbar
from tkinter import filedialog
from PIL import Image, ImageTk
import os
import openpyxl
import cv2
import imageio
from math import sqrt


# 全局变量
total_folder_num = 0
current_main_folder_index = 0
current_sub_folder_index = 0
info_file_path = ""
screenshot_images = []
source_images = []
main_folder_path = ""
sub_folders = []
source_label = None 


def template_match(origin_path,template_path):
    threshold = 0.75
    image = imageio.imread(origin_path)
    template = imageio.imread(template_path)
    # !!!使用cv2默认的imread会无法读取gif等格式 这里用imageio
    # 将图像和模板都转换为灰度
    imageGray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
    templateGray = cv2.cvtColor(template, cv2.COLOR_RGB2GRAY)
    if template.shape[0] > image.shape[0] or template.shape[1] > image.shape[1]:
        return False, -1

    result = cv2.matchTemplate(imageGray, templateGray, cv2.TM_CCOEFF_NORMED)
    (minVal, maxVal, minLoc, maxLoc) = cv2.minMaxLoc(result)
    print(maxVal)
    color = (0,0,255)
    if maxVal >= threshold:
        color = (0,255,0)
    if maxVal == 1.0:
        maxVal = 0#剔除无效结果
    (startX, startY) = maxLoc
    endX = startX + template.shape[1]
    endY = startY + template.shape[0]
    center_x = (startX+endX)/2
    center_y = (startY+endY)/2
    dist_offset = sqrt((center_x - image.shape[1]/2)*(center_x - image.shape[1]/2)+(center_y - image.shape[0]/2)*(center_y - image.shape[0]/2))
    return True if maxVal>threshold else False, dist_offset


def open_folder():
    global main_folder_path,total_folder_num,current_sub_folder_index,current_main_folder_index
    folder_path = filedialog.askdirectory()  # 打开文件夹选择对话框
    if folder_path:
        main_folder_path = folder_path
        main_folders = [f for f in os.listdir(main_folder_path) if os.path.isdir(os.path.join(main_folder_path, f))]
        total_folder_num = len(os.listdir(main_folder_path))
        print(os.listdir(main_folder_path))
        print(total_folder_num)
        total_folder_num-=1
        for widget in checkboxes_frame.winfo_children():
            widget.destroy()
        current_main_folder_index = 0
        current_sub_folder_index = 0
        load_images()

class fuck_pussy():
    name = ''
    fucker = ''
    def __init__(self,name,fucker) -> None:
        self.name = name
        self.fucker = fucker
        pass
    
    def fuck(self):
        print(f'小比{self.name}挨测，{self.fucker}的东西大不大，爽不爽')


def openpic(path):
    os.system('start '+path)

def load_images():
    global current_main_folder_index, current_sub_folder_index, info_file_path, screenshot_images, main_folder_path, sub_folders, source_label

    main_folder_name = str(current_main_folder_index)

    sub_folders = [f for f in os.listdir(os.path.join(main_folder_path, main_folder_name)) if os.path.isdir(os.path.join(main_folder_path, main_folder_name, f))]
    ttk.Label()

    offset = []
    index = []
    for idx in range(len(sub_folders)):
        pth = os.path.join(main_folder_path,main_folder_name,str(idx))
        files = os.listdir(pth)
        screenshot = ''
        source = ''
        for f in files:
            if f[:6]=='source':
                source = f
            elif f[:10] == 'screenshot':
                screenshot = f
        match,oset = template_match(os.path.join(pth,screenshot),os.path.join(pth,source))
        if match:
            offset.append(oset)
            index.append(idx)

    best_optn = index[offset.index(min(offset))] if len(offset)!=0 else None
    print(offset)
    print(best_optn)

    if current_sub_folder_index < len(sub_folders):
        sub_folder_name = sub_folders[current_sub_folder_index]

        # 拼接当前子文件夹路径
        current_folder_path = os.path.join(main_folder_path, main_folder_name, sub_folder_name)
    
        # 拼接infos.xlsx文件路径
        info_file_path = os.path.join(main_folder_path, main_folder_name, "infos.xlsx")

        # 打开infos.xlsx
        wb = openpyxl.load_workbook(info_file_path)
        ws = wb.active

        global source_img
        screenshot_images = []

        open_bigshot_buttons = []
        open_source_buttons = []
        open_screenshot_buttons = []
        
        # 获取当前子文件夹内的所有screenshot.png
        for i in range(len(sub_folders)):
            source_image_path = ""
            fnames = os.listdir(os.path.join(main_folder_path,main_folder_name,sub_folders[i]))
            current_folder_path = os.path.join(main_folder_path,main_folder_name,sub_folders[i])
            source_fname = ''
            for f in fnames:
                if f[:6] == "source":
                    source_image_path = os.path.join(current_folder_path, f)
                    source_fname = f
                    break
            screenshot_image_path = os.path.join(main_folder_path, main_folder_name, sub_folders[i], "screenshot.png")
            screenshot_img = Image.open(screenshot_image_path)
            screenshot_img.thumbnail((200, 200))
            screenshot_img = ImageTk.PhotoImage(screenshot_img)
            screenshot_images.append(screenshot_img)  # 将 screenshot 图片添加到列表

            source_img = Image.open(source_image_path)
            source_img.thumbnail((200,200))
            source_img = ImageTk.PhotoImage(source_img)
            source_images.append(source_img)

            # 创建标签和复选框
            screenshot_label = ttk.Label(checkboxes_frame, image=screenshot_img)
            source_label = ttk.Label(checkboxes_frame,image=source_img)
            valid_checkbox = ttk.Checkbutton(checkboxes_frame, text="Valid")
            ad_checkbox = ttk.Checkbutton(checkboxes_frame, text="Ad")
            #open_full_button = ttk.Button(checkboxes_frame,text='打开大图', command= lambda:show_img(os.path.join(current_folder_path,'bigshot.png')))
            #open_source_button = ttk.Button(checkboxes_frame,text='打开原图', command = lambda:show_img(os.path.join(main_folder_path,main_folder_name,str(i),source_fname)))
            #open_screenshot_button = ttk.Button(checkboxes_frame,text='打开截图', command = lambda:show_img(screenshot_image_path))
            open_full_button = ttk.Button(checkboxes_frame,text='打开大图', command= lambda f=os.path.join(current_folder_path,'bigshot.png'):show_img(f))
            open_source_button = ttk.Button(checkboxes_frame,text='打开原图', command = lambda f=os.path.join(current_folder_path,source_fname):show_img(f))
            open_screenshot_button = ttk.Button(checkboxes_frame,text='打开截图', command = lambda f=os.path.join(current_folder_path,'screenshot.png'):show_img(f))
            #similarity_text = ttk.Label(checkboxes_frame,text=)
            # 初始化复选框
            valid_checkbox.state(['!alternate'])
            ad_checkbox.state(['!alternate'])

            # 显示图片和复选框
            source_label.grid(row=i,column=1,padx=5,pady=5)
            screenshot_label.grid(row=i, column=0, padx=5, pady=5)
            valid_checkbox.grid(row=i, column=2, padx=5, pady=5)
            ad_checkbox.grid(row=i, column=3, padx=5, pady=5)
            open_full_button.grid(row=i,column=4,padx=5,pady=5)
            open_source_button.grid(row=i,column=5,padx=5,pady=5)
            open_screenshot_button.grid(row=i,column=6,padx=5,pady=5)

            if i == best_optn:
                #valid_checkbox.state(['selected'])
                recommend_label = ttk.Label(checkboxes_frame,text="MATCH",font=('微软雅黑',20),foreground='green')
                recommend_label.grid(row=i,column=7,padx=5,pady=5)
            # 从infos.xlsx读取并设置复选框状态
            valid_value = ws.cell(row=i+2, column=5).value
            ad_value = ws.cell(row=i+2, column=6).value
            if valid_value == 1:
                valid_checkbox.state(['selected'])
            if ad_value == 1:
                ad_checkbox.state(['selected'])

        # 释放infos.xlsx文件
        wb.close()
        update_scrollregion()

def save_changes():
    global total_folder_num,current_main_folder_index
    global info_file_path
    if not info_file_path:
        return

    wb = openpyxl.load_workbook(info_file_path)
    ws = wb.active

    # 更新infos.xlsx中的数据
    row1 = 1
    row2 = 1
    for i, widget in enumerate(checkboxes_frame.winfo_children()):
        if isinstance(widget, ttk.Checkbutton):
            col = 5
            if widget.cget("text") == "Ad":
                col = 6
                row2 += 1
                value = 1 if widget.instate(["selected"]) else 0
                ws.cell(row=row2, column=col, value=value)
            elif widget.cget("text") == "Valid":
                row1 += 1
                value = 1 if widget.instate(["selected"]) else 0
                ws.cell(row=row1, column=col, value=value)


    wb.save(info_file_path)
    wb.close()

    # 清除旧内容
    for widget in checkboxes_frame.winfo_children():
        widget.destroy()
    source_img.__del__()
    if current_main_folder_index == total_folder_num-1:
        tk.messagebox.showinfo(title='提示',message='已完成最后一个文件组的处理！')
        return
    current_main_folder_index+=1
    load_images()
    update_scrollregion()


def show_img(img_path):

    os.system(f'''\"{img_path}\"''')
    print(f'''\"{img_path}\"''')

def back():
    global current_main_folder_index
    if current_main_folder_index==0:
        tk.messagebox.showinfo(title='提示',message='当前已是第一个图片组')
        return
    current_main_folder_index-=1
    print(current_main_folder_index)
    # 清除旧内容
    for widget in checkboxes_frame.winfo_children():
        widget.destroy()
    source_img.__del__()
    load_images()
    update_scrollregion()

def on_mousewheel(event):
    if event.delta > 0:
        canvas.yview_scroll(-2, "units")  # 向上滚动
    elif event.delta < 0:
        canvas.yview_scroll(2, "units")  # 向下滚动

def update_scrollregion():
    window_width,window_height=root.winfo_width(),root.winfo_height()
    canvas.update_idletasks()  # 更新canvas
    root.geometry(str(window_width)+'x'+str(window_height+1))
    root.geometry(str(window_width)+'x'+str(window_height))
    canvas.yview_scroll(-1000,'units')
    canvas.config(scrollregion=canvas.bbox("all"))  # 重新设置滚动区域

if __name__ == '__main__':

    root = tk.Tk()
    root.title("img tagger")
    root.geometry('1280x720')

    main_folder_path = ""
    main_folders = []

    source_label = ttk.Label(root)
    source_label.pack(padx=20, pady=20)

    canvas = tk.Canvas(root)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar = ttk.Scrollbar(root, orient=tk.VERTICAL, command=canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    canvas.configure(yscrollcommand=scrollbar.set)

    checkboxes_frame = ttk.Frame(canvas)
    canvas.create_window((0, 0), window=checkboxes_frame, anchor=tk.NW)

    open_folder_button = ttk.Button(root, text="打开文件夹", command=open_folder)
    open_folder_button.pack(padx=20, pady=5)

    save_button = ttk.Button(root, text="保存修改", command=save_changes)
    save_button.pack(padx=20, pady=5)

    back_button = ttk.Button(root,text='上个图片组',command=back)
    back_button.pack(padx=20, pady=5)

    root.bind("<Return>", lambda event=None: save_changes())
    root.bind("<BackSpace>",lambda event = None: back())
    root.bind("<MouseWheel>", on_mousewheel)
    canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    root.mainloop()
